# -*- coding: utf-8 -*-
# from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook
import os


def readXlsxFile(path):
    # 打开文件
    file = load_workbook(filename=path)
    # 获取excel页签名称
    sheets = file.get_sheet_names()
    # print(file.get_sheet_names())
    dic = {}
    for sheetname in sheets:
        sheet_content = []
        # 获取一个表格(表格的序号是从0开始的)
        sheet = file.get_sheet_by_name(sheetname)
        # 获取最大行数 列数 标题
        # print(sheet.max_row)
        # print(sheet.max_column)
        # print(sheet.title)
        for linNUM in range(1, sheet.max_row + 1):
            # print(linNUM)
            # 定义坐标如果有值把坐标添加到列表
            # positions = []
            line_content = []
            for colNUM in range(1, sheet.max_column + 1):

                # sheet.cell获取指定位置的数据(起始位置为1,1)
                value = sheet.cell(row=linNUM, column=colNUM)
                # sheet.cell(row=linNUM, column=colNUM).value 获取指定位置的值
                if value.value != None:
                    # 把每个位置不为空的内容添加到行列表中
                    line_content.append(value.value)
                    # position = '%s,%s' % (linNUM, colNUM)
                    # positions.append(position)
            # 将每行内容添加到每页的内容中
            sheet_content.append(line_content)
        # 把每页的内容添加到字典
        dic[sheetname] = sheet_content
    return dic
            # print('数据位置:', positions)
            # print('返回结果:', line_content)

# 执行读取excel文档的动作
def runreadxlsx():

    path = os.path.join(os.getcwd(), 'file\浙商测试云运维月报20180901-20180931.xlsx',)
    # path = os.path.join(os.getcwd(), 'file\浙商云运维月报20170205-20170405.xlsx',)
    # path = os.path.join(os.getcwd(), 'file\\test.xlsx',)
    dic = readXlsxFile(path)
    print(dic)

runreadxlsx()